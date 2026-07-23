"""
Royal Commission Clinics (Jubail) - Quarterly Insurance Claims Aggregator
Streamlit app that reads the "SUM" sheet from 3 monthly Excel files (7 fixed
clinics each), aggregates Status and Rejection-Reason tables across all
clinics/months, validates totals, and exports a formatted Excel summary.
"""

import io
from datetime import datetime

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------

INSURANCE_COMPANIES = [
    "Bupa",
    "Tawuniya",
    "Medgulf",
    "Malath",
    "SAICO",
    "Gulf Union",
    "Alrajhi Takaful",
    "Gulf Insurance Group – GIG",
    "TCS",
    "GLOBMED",
    "NEXTCARE",
    "ARABIAN SHIELD",
]

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

SUM_SHEET_NAME = "SUM"

CLINICS = [
    {"name": "Al Osrah Medical Center", "start": 4, "end": 22},
    {"name": "Al Dafi First Aid Medical Center", "start": 26, "end": 44},
    {"name": "Al Howilat First Aid Medical Center", "start": 48, "end": 65},
    {"name": "Al Farouq First Aid Medical Center", "start": 69, "end": 87},
    {"name": "Jalmoud First Aid Medical Center", "start": 91, "end": 109},
    {"name": "Royal Commission Hospital", "start": 113, "end": 131},
    {"name": "Ras Al Khair First Aid Medical Center", "start": 135, "end": 153},
]

STATUS_COL = {"name": 1, "cases": 2, "amount": 3}   # A, B, C
REASON_COL = {"name": 7, "cases": 8, "amount": 9}   # G, H, I ("Total Reasons" table)


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def safe_number(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).replace(",", "").strip()
        if cleaned == "":
            return 0.0
        return float(cleaned)
    except (TypeError, ValueError):
        return 0.0


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def is_total_or_blank(text):
    normalized = text.strip().lower()
    if normalized == "":
        return True
    return "total" in normalized


def extract_sum_sheet(file_obj, month_name):
    wb = openpyxl.load_workbook(file_obj, data_only=True)

    if SUM_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"The sheet '{SUM_SHEET_NAME}' was not found in the file for {month_name}."
        )

    ws = wb[SUM_SHEET_NAME]

    status_rows = []
    reason_rows = []

    for clinic in CLINICS:
        for row in range(clinic["start"], clinic["end"] + 1):
            status_val = clean_text(ws.cell(row=row, column=STATUS_COL["name"]).value)
            if status_val and not is_total_or_blank(status_val):
                status_rows.append(
                    {
                        "Status": status_val,
                        "Cases": safe_number(ws.cell(row=row, column=STATUS_COL["cases"]).value),
                        "NetAmount+Vat": safe_number(
                            ws.cell(row=row, column=STATUS_COL["amount"]).value
                        ),
                        "Clinic": clinic["name"],
                        "Month": month_name,
                    }
                )

            reason_val = clean_text(ws.cell(row=row, column=REASON_COL["name"]).value)
            if reason_val and not is_total_or_blank(reason_val):
                reason_rows.append(
                    {
                        "Reason": reason_val,
                        "Cases": safe_number(ws.cell(row=row, column=REASON_COL["cases"]).value),
                        "NetAmount+Vat": safe_number(
                            ws.cell(row=row, column=REASON_COL["amount"]).value
                        ),
                        "Clinic": clinic["name"],
                        "Month": month_name,
                    }
                )

    if not status_rows and not reason_rows:
        raise ValueError(f"No data could be read from the '{SUM_SHEET_NAME}' sheet for {month_name}.")

    return status_rows, reason_rows


NAME_COLUMN_SYNONYMS = {
    "Status": {"status", "statuses", "claim status"},
    "Reason": {"reason", "reasons", "rejection reason", "rejection reasons"},
}


def _normalize_table(rows, name_col):
    expected_cols = [name_col, "Cases", "NetAmount+Vat"]
    df = pd.DataFrame(rows)

    synonyms = NAME_COLUMN_SYNONYMS.get(name_col, {name_col.lower()})
    rename_map = {}
    for col in df.columns:
        cleaned = str(col).strip()
        if cleaned.lower() in synonyms:
            rename_map[col] = name_col
        elif cleaned != col:
            rename_map[col] = cleaned
    if rename_map:
        df = df.rename(columns=rename_map)

    for col in expected_cols:
        if col not in df.columns:
            df[col] = 0.0 if col != name_col else ""

    df[name_col] = df[name_col].fillna("").astype(str).str.strip()

    for numeric_col in ["Cases", "NetAmount+Vat"]:
        df[numeric_col] = pd.to_numeric(df[numeric_col], errors="coerce").fillna(0)

    df = df[df[name_col] != ""]
    return df[expected_cols]


def aggregate(status_rows, reason_rows):
    status_df = _normalize_table(status_rows, "Status")
    reason_df = _normalize_table(reason_rows, "Reason")

    status_summary = (
        status_df.groupby("Status", as_index=False)[["Cases", "NetAmount+Vat"]]
        .sum()
        .sort_values("Status")
        .reset_index(drop=True)
    )
    reason_summary = (
        reason_df.groupby("Reason", as_index=False)[["Cases", "NetAmount+Vat"]]
        .sum()
        .sort_values("Reason")
        .reset_index(drop=True)
    )

    status_summary["Cases"] = status_summary["Cases"].round(0).astype(int)
    reason_summary["Cases"] = reason_summary["Cases"].round(0).astype(int)

    return status_summary, reason_summary


def build_download_workbook(insurance_company, status_summary, reason_summary):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quarterly Summary"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=13)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    def write_title(row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = title_font
        return row + 1

    def write_table(start_row, title, df, name_col_header):
        row = write_title(start_row, title)
        headers = [name_col_header, "Total Cases", "Total NetAmount+Vat (SAR)"]
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        row += 1
        for _, r in df.iterrows():
            ws.cell(row=row, column=1, value=r.iloc[0]).border = border
            cases_cell = ws.cell(row=row, column=2, value=int(r["Cases"]))
            cases_cell.border = border
            cases_cell.number_format = "#,##0"
            amt_cell = ws.cell(row=row, column=3, value=float(r["NetAmount+Vat"]))
            amt_cell.border = border
            amt_cell.number_format = '#,##0.00 "SAR"'
            row += 1

        total_cases = int(df["Cases"].sum())
        total_amount = float(df["NetAmount+Vat"].sum())
        tot_label = ws.cell(row=row, column=1, value="TOTAL")
        tot_label.font = Font(bold=True)
        tot_label.border = border
        tot_cases = ws.cell(row=row, column=2, value=total_cases)
        tot_cases.font = Font(bold=True)
        tot_cases.number_format = "#,##0"
        tot_cases.border = border
        tot_amt = ws.cell(row=row, column=3, value=total_amount)
        tot_amt.font = Font(bold=True)
        tot_amt.number_format = '#,##0.00 "SAR"'
        tot_amt.border = border
        return row + 2

    row = 1
    header_cell = ws.cell(row=row, column=1, value=f"Insurance Company: {insurance_company}")
    header_cell.font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    row += 2

    row = write_table(row, "Table 1 — Total Status Summary", status_summary, "Status")
    row = write_table(row, "Table 2 — Total Rejection Reasons Summary", reason_summary, "Reason")

    for col_idx in range(1, 4):
        letter = get_column_letter(col_idx)
        max_len = 12
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --------------------------------------------------------------------------
# Streamlit UI
# --------------------------------------------------------------------------

st.set_page_config(
    page_title="Royal Commission Jubail — Quarterly Claims Aggregator",
    layout="wide",
)

CORPORATE_CSS = """
<style>
/* ---- Font: Times New Roman everywhere ---- */
html, body, [class*="css"], .stApp, .stMarkdown, .stText,
.stSelectbox, .stFileUploader, .stMultiSelect, .stButton, .stDownloadButton,
table.corporate-table, input, textarea, select, button, label, p, span, div {
    font-family: "Times New Roman", Times, serif !important;
}

/* ---- Page background ---- */
.stApp {
    background-color: #DDDEEB;
    color: #111111;
}

/* ---- Headings ---- */
h1, h2, h3 {
    color: #2C3E50 !important;
    font-weight: 700 !important;
    font-family: "Times New Roman", Times, serif !important;
}

/* ---- Field labels ---- */
.stSelectbox label, .stFileUploader label, .stMultiSelect label,
div[data-testid="stWidgetLabel"] p {
    color: #111111 !important;
    font-weight: 600 !important;
}

/* ---- Selectboxes (Insurance Company, Months) ---- */
div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
div[data-baseweb="select"] > div,
div[data-baseweb="select"],
div[role="combobox"] {
    background-color: #FFFFFF !important;
    color: #000000 !important;
    border: 1px solid #2C3E50 !important;
    border-radius: 6px !important;
}
div[data-testid="stSelectbox"] div[data-baseweb="select"] *,
div[data-baseweb="select"] * {
    color: #000000 !important;
    fill: #000000 !important;
    font-family: "Times New Roman", Times, serif !important;
}
div[data-baseweb="select"] span {
    color: #000000 !important;
    font-weight: 500 !important;
}

/* Dropdown popover menu + options */
ul[data-baseweb="menu"], div[data-baseweb="popover"] ul,
ul[role="listbox"], li[role="option"] {
    background-color: #FFFFFF !important;
    border: 1px solid #a5a9c9 !important;
    border-radius: 6px !important;
}
ul[data-baseweb="menu"] li, div[data-baseweb="popover"] li,
li[role="option"] {
    background-color: #FFFFFF !important;
    color: #000000 !important;
}
ul[data-baseweb="menu"] li *, div[data-baseweb="popover"] li *,
li[role="option"] * {
    color: #000000 !important;
}
ul[data-baseweb="menu"] li:hover, div[data-baseweb="popover"] li:hover,
li[role="option"]:hover {
    background-color: #D0D3E5 !important;
    color: #000000 !important;
}

/* ---- File uploaders styling ---- */
section[data-testid="stFileUploaderDropzone"],
div[data-testid="stFileUploader"] section {
    background-color: #D7D8E0 !important;
    border: 1px dashed #2C3E50 !important;
    border-radius: 10px !important;
}

/* Clean up uploaded file box and show built-in delete button properly */
div[data-testid="stFileUploaderFile"] {
    background-color: #FFFFFF !important;
    color: #111111 !important;
    border: 1px solid #2C3E50 !important;
    border-radius: 6px !important;
}

/* Ensure Native Streamlit Delete Button is visible and clean */
div[data-testid="stFileUploaderFile"] button {
    background-color: transparent !important;
    border: none !important;
    color: #2C3E50 !important;
}
div[data-testid="stFileUploaderFile"] button:hover {
    color: #C0392B !important;
    background-color: rgba(192, 57, 43, 0.1) !important;
    border-radius: 50% !important;
}

/* ---- Card wrapper for HTML tables ---- */
.corporate-card {
    background-color: #E9EAF3;
    border: 1px solid #B6BAD9;
    border-radius: 10px;
    padding: 16px 18px;
    margin-bottom: 18px;
}
.corporate-card h4 {
    color: #2C3E50;
    margin-top: 0;
    font-weight: 700;
}

/* ---- Data tables ---- */
table.corporate-table {
    width: 100%;
    border-collapse: collapse;
    background-color: #FFFFFF;
    border-radius: 8px;
    overflow: hidden;
}
table.corporate-table th {
    background-color: #2C3E50;
    color: #FFFFFF !important;
    padding: 10px 12px;
    text-align: left;
    font-weight: 600;
    border: 1px solid #B6BAD9;
}
table.corporate-table td {
    padding: 8px 12px;
    color: #111111;
    border: 1px solid #B6BAD9;
}
table.corporate-table tr:nth-child(even) td {
    background-color: #F1F2F8;
}
table.corporate-table tr:last-child td {
    font-weight: 700;
    background-color: #DDDEEB;
}

/* ---- Validation boxes ---- */
.validation-box {
    padding: 14px 18px;
    border-radius: 10px;
    font-weight: 600;
    margin-top: 8px;
    border: 1px solid #B6BAD9;
}
.validation-success {
    background-color: #b8d4ba !important;
    color: #111111;
}
.validation-error {
    background-color: #F3C6C6;
    color: #6B1414;
}

/* ---- Alert notifications ---- */
[data-testid="stAlert"] {
    background-color: #b8d4ba !important;
    color: #111111 !important;
    border: 1px solid #B6BAD9 !important;
}
[data-testid="stAlert"] * {
    color: #111111 !important;
}

/* ---- Buttons (Process + Download) ---- */
div.stDownloadButton,
div.stButton {
    margin-top: 20px;
}
div.stDownloadButton > button,
div.stButton > button {
    background-color: #2C3E50;
    color: #FFFFFF !important;
    border: 1px solid #2C3E50;
    border-radius: 8px;
    padding: 0.5em 1.3em;
    font-weight: 600;
    transition: all 0.2s ease-in-out;
}
div.stDownloadButton > button:hover,
div.stButton > button:hover {
    background-color: #3E5670;
    border-color: #3E5670;
    color: #FFFFFF !important;
    box-shadow: 0 2px 8px rgba(44, 62, 80, 0.35);
}
</style>
"""
st.markdown(CORPORATE_CSS, unsafe_allow_html=True)


st.title("📊 Quarterly Insurance Claims Aggregator")
st.caption("Royal Commission clinics — Jubail Industrial")

if "results" not in st.session_state:
    st.session_state["results"] = None

insurance_company = st.selectbox("Insurance Company", INSURANCE_COMPANIES)

st.markdown("Upload the 3 monthly files for this quarter")
cols = st.columns(3)
uploads = []
for i, col in enumerate(cols, start=1):
    with col:
        st.markdown(f"**File {i}**")
        f = st.file_uploader(
            f"File {i}", type=["xlsx", "xls"], key=f"file_{i}", label_visibility="collapsed"
        )
        m = st.selectbox(f"Month for File {i}", MONTH_NAMES, index=(i - 1) % 12, key=f"month_{i}")
        uploads.append({"file": f, "month": m})

process_clicked = st.button("Process Quarter", type="primary")

if process_clicked:
    if any(u["file"] is None for u in uploads):
        st.warning("Please upload all 3 files before processing.")
    elif len({u["month"] for u in uploads}) != 3:
        st.warning("Please select 3 different months (one per file).")
    else:
        all_status_rows = []
        all_reason_rows = []
        error_occurred = False
        with st.spinner("Reading and aggregating data..."):
            for u in uploads:
                try:
                    s_rows, r_rows = extract_sum_sheet(u["file"], u["month"])
                    all_status_rows.extend(s_rows)
                    all_reason_rows.extend(r_rows)
                except ValueError as e:
                    st.error(str(e))
                    error_occurred = True
                except Exception as e:
                    st.error(f"Unexpected error reading file for {u['month']}: {e}")
                    error_occurred = True

        if not error_occurred:
            status_summary, reason_summary = aggregate(all_status_rows, all_reason_rows)
            st.session_state["results"] = {
                "insurance_company": insurance_company,
                "status_summary": status_summary,
                "reason_summary": reason_summary,
                "months": [u["month"] for u in uploads],
            }
            st.success("Quarter processed successfully.")

results = st.session_state["results"]

if results:
    status_summary = results["status_summary"]
    reason_summary = results["reason_summary"]

    st.markdown("---")
    st.markdown(
        f"### Results — **{results['insurance_company']}** "
        f"({', '.join(results['months'])})"
    )

    total_cases_status = int(status_summary["Cases"].sum())
    total_cases_reasons = int(reason_summary["Cases"].sum())

    def with_total_row(df, name_col):
        total_row = pd.DataFrame(
            [{name_col: "TOTAL", "Cases": df["Cases"].sum(), "NetAmount+Vat": df["NetAmount+Vat"].sum()}]
        )
        return pd.concat([df, total_row], ignore_index=True)

    def render_corporate_table(df, name_col, card_title):
        html_table = df.to_html(index=False, classes="corporate-table", border=0, escape=False)
        st.markdown(
            f"""<div class="corporate-card">
                    <h4>{card_title}</h4>
                    {html_table}
                </div>""",
            unsafe_allow_html=True,
        )

    c1, c2 = st.columns(2)
    with c1:
        display_status = with_total_row(status_summary, "Status")
        display_status["Cases"] = display_status["Cases"].map("{:,.0f}".format)
        display_status["NetAmount+Vat"] = display_status["NetAmount+Vat"].map(
            lambda x: f"{x:,.2f} SAR"
        )
        render_corporate_table(display_status, "Status", "Table 1 — Total Status Summary")

    with c2:
        display_reasons = with_total_row(reason_summary, "Reason")
        display_reasons["Cases"] = display_reasons["Cases"].map("{:,.0f}".format)
        display_reasons["NetAmount+Vat"] = display_reasons["NetAmount+Vat"].map(
            lambda x: f"{x:,.2f} SAR"
        )
        render_corporate_table(display_reasons, "Reason", "Table 2 — Total Rejection Reasons Summary")

    st.markdown("### Validation Check")
    if total_cases_status == total_cases_reasons:
        st.markdown(
            f"""<div class="validation-box validation-success">
                    Balanced: Total Cases (Status) = {total_cases_status:,}
                    = Total Cases (Reasons) = {total_cases_reasons:,}
                </div>""",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f"""<div class="validation-box validation-error">
                    ⚠️ Mismatch detected — Total Cases (Status) = {total_cases_status:,}
                    vs Total Cases (Reasons) = {total_cases_reasons:,}
                    (difference of {abs(total_cases_status - total_cases_reasons):,})
                </div>""",
            unsafe_allow_html=True,
        )

    excel_buffer = build_download_workbook(
        results["insurance_company"], status_summary, reason_summary
    )
    file_name = (
        f"{results['insurance_company'].replace(' ', '_')}_"
        f"Quarterly_Claims_Summary.xlsx"
    )
    st.download_button(
        label="Download Excel Summary",
        data=excel_buffer,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("Upload 3 files, select their months, and click **Process Quarter** to see results.")

# --------------------------------------------------------------------------
# Footer (حقوق الملكية)
# --------------------------------------------------------------------------
st.markdown("---")
st.markdown(
    """
    <div style='text-align: center; padding: 10px 0;'>
        <p style='color: #444444 !important; font-size: 13px !important; font-family: "Times New Roman", Times, serif !important; margin: 0;'>
            Developed by <b>Raghad Alqarni</b> | All Rights Reserved © 2026
        </p>
    </div>
    """,
    unsafe_allow_html=True
)
